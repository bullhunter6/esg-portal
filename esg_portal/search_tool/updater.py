from openpyxl import Workbook, load_workbook
from search_tool.score_extractors.snp import snp_score
from search_tool.score_extractors.sustainalytics import sustainalytics_score
from search_tool.score_extractors.iss import iss_score
from search_tool.score_extractors.lseg import lseg_score
from search_tool.score_extractors.msci import msci_score
from search_tool.score_extractors.cdp import cdp_score


def fetch_scores(company_name, source_specific_names):
    """
    Fetch ESG scores from various sources using the company name and source-specific names.
    """
    scores = {}
    details = {}

    def get_safe_name(source):
        return source_specific_names.get(source, company_name).strip()

    # Fetch S&P score
    try:
        s_and_p_name = get_safe_name("S&P")
        if s_and_p_name:
            snp_data = snp_score(s_and_p_name)
            scores["S&P"] = snp_data.get("esg_score", '-') if isinstance(snp_data, dict) else '-'
            details["S&P"] = snp_data
        else:
            print(f"Skipping S&P fetch for {company_name}: No valid name")
            scores["S&P"] = '-'
    except Exception as e:
        print(f"Error fetching S&P score for {company_name}: {e}")
        scores["S&P"] = '-'

    # Fetch Sustainalytics score
    try:
        sustainalytics_name = get_safe_name("Sustainalytics")
        if sustainalytics_name:
            sustainalytics_data = sustainalytics_score(sustainalytics_name)
            scores["Sustainalytics"] = sustainalytics_data.get("esg_score", '-') if isinstance(sustainalytics_data, dict) else '-'
            details["Sustainalytics"] = sustainalytics_data
        else:
            print(f"Skipping Sustainalytics fetch for {company_name}: No valid name")
            scores["Sustainalytics"] = '-'
    except Exception as e:
        print(f"Error fetching Sustainalytics score for {company_name}: {e}")
        scores["Sustainalytics"] = '-'

    # Fetch ISS score
    try:
        iss_name = get_safe_name("ISS")
        if iss_name:
            iss_data = iss_score(iss_name)
            scores["ISS"] = iss_data[0].get("oekomRating", '-') if isinstance(iss_data, list) and len(iss_data) > 0 else '-'
            details["ISS"] = iss_data
        else:
            print(f"Skipping ISS fetch for {company_name}: No valid name")
            scores["ISS"] = '-'
    except Exception as e:
        print(f"Error fetching ISS score for {company_name}: {e}")
        scores["ISS"] = '-'

    # Fetch LSEG score
    try:
        lseg_name = get_safe_name("LSEG")
        if lseg_name:
            lseg_data = lseg_score(lseg_name)
            scores["LSEG"] = lseg_data.get("TR.TRESG", '-') if isinstance(lseg_data, dict) else '-'
            details["LSEG"] = lseg_data
        else:
            print(f"Skipping LSEG fetch for {company_name}: No valid name")
            scores["LSEG"] = '-'
    except Exception as e:
        print(f"Error fetching LSEG score for {company_name}: {e}")
        scores["LSEG"] = '-'

    # Fetch MSCI score
    try:
        msci_name = get_safe_name("MSCI")
        if msci_name:
            msci_data = msci_score(msci_name)
            scores["MSCI"] = msci_data.get("ESG Rating", '-') if isinstance(msci_data, dict) else '-'
            details["MSCI"] = msci_data
        else:
            print(f"Skipping MSCI fetch for {company_name}: No valid name")
            scores["MSCI"] = '-'
    except Exception as e:
        print(f"Error fetching MSCI score for {company_name}: {e}")
        scores["MSCI"] = '-'

    return company_name, scores, details




def worker(args):
    return fetch_scores(*args)


def update_excel(input_file, results, output_file):
    """
    Update the Excel file with fetched scores and details.
    """
    try:
        workbook = load_workbook(input_file)
    except FileNotFoundError:
        workbook = Workbook()

    # Create Summary sheet if it doesn't exist
    if "Summary" not in workbook.sheetnames:
        summary_sheet = workbook.create_sheet("Summary")
        summary_sheet.append([
            "Company Name", "S&P", "Sustainalytics", "ISS", "LSEG", "MSCI"
        ])
    else:
        summary_sheet = workbook["Summary"]

    # Create detailed sheets for each source
    details_sheets = {}
    for source in ["S&P", "Sustainalytics", "ISS", "LSEG", "MSCI"]:
        if source not in workbook.sheetnames:
            sheet = workbook.create_sheet(source)
            details_sheets[source] = sheet
            initialize_columns_for_source(sheet, source)
        else:
            details_sheets[source] = workbook[source]

    # Populate summary sheet
    for company_name, scores, details in results:
        summary_sheet.append([
            company_name,
            scores.get("S&P", '-'),
            scores.get("Sustainalytics", '-'),
            scores.get("ISS", '-'),
            scores.get("LSEG", '-'),
            scores.get("MSCI", '-')
        ])

        # Append details to detailed sheets
        append_details_to_sheets(details_sheets, company_name, details)

    # Save the workbook
    workbook.save(output_file)
    print(f"Updated Excel file saved at {output_file}.")


def initialize_columns_for_source(sheet, source):
    """
    Initialize columns for each source's detailed sheet.
    """
    if source == "S&P":
        sheet.append(["Company Name", "ESG Score", "Details"])
    elif source == "Sustainalytics":
        sheet.append(["Company Name", "ESG Score", "Details"])
    elif source == "ISS":
        sheet.append(["Company Name", "Oekom Rating", "Details"])
    elif source == "LSEG":
        sheet.append(["Company Name", "TR.TRESG", "Details"])
    elif source == "MSCI":
        sheet.append(["Company Name", "ESG Rating", "Details"])

def append_details_to_sheets(details_sheets, company_name, details):
    """
    Append detailed data to respective sheets.
    """
    for source, data in details.items():
        source_sheet = details_sheets.get(source)
        if source == "ISS" and isinstance(data, dict):
            if source_sheet.max_row == 1:
                source_sheet.append(["Company Name"] + list(data.keys()))
            source_sheet.append([company_name] + list(data.values()))
        elif source != "ISS" and isinstance(data, dict):
            if source_sheet.max_row == 1:
                source_sheet.append(["Company Name"] + list(data.keys()))
            source_sheet.append([company_name] + list(data.values()))



