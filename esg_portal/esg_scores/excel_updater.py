"""
Excel updater for ESG scores
"""
import os
import pandas as pd
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from flask import current_app
import concurrent.futures
from functools import lru_cache

# Import score extractors
from esg_portal.search_tool.score_extractors.snp import snp_score
from esg_portal.search_tool.score_extractors.sustainalytics import sustainalytics_score
from esg_portal.search_tool.score_extractors.iss import iss_score
from esg_portal.search_tool.score_extractors.lseg import lseg_score
from esg_portal.search_tool.score_extractors.msci import msci_score
from esg_portal.search_tool.score_extractors.cdp import cdp_score

# Cache for API calls
@lru_cache(maxsize=1000)
def cached_snp_score(name):
    return snp_score(name)

@lru_cache(maxsize=1000)
def cached_sustainalytics_score(name):
    return sustainalytics_score(name)

@lru_cache(maxsize=1000)
def cached_iss_score(name):
    return iss_score(name)

@lru_cache(maxsize=1000)
def cached_lseg_score(name):
    return lseg_score(name)

@lru_cache(maxsize=1000)
def cached_msci_score(name):
    return msci_score(name)

def fetch_single_source(source, name):
    """Fetch score from a single source with caching"""
    try:
        if not name:
            return '-', {}
            
        if source == "S&P":
            data = cached_snp_score(name)
            score_key = "esg_score"
        elif source == "Sustainalytics":
            data = cached_sustainalytics_score(name)
            score_key = "esg_score"
        elif source == "ISS":
            data = cached_iss_score(name)
            score_key = "oekomRating"
        elif source == "LSEG":
            data = cached_lseg_score(name)
            score_key = "TR.TRESG"
        elif source == "MSCI":
            data = cached_msci_score(name)
            score_key = "ESG Rating"
        else:
            return '-', {}
            
        # Handle ISS data which comes as a list
        if source == "ISS" and isinstance(data, list) and len(data) > 0:
            data = data[0]
            
        # Extract score based on the source-specific key
        if isinstance(data, dict):
            score = data.get(score_key, '-')
            # For LSEG, if TR.TRESG is not found, try TR.TRESG.Score
            if source == "LSEG" and score == '-':
                score = data.get("TR.TRESG.Score", '-')
            # For MSCI, if ESG Rating is not found, try Rating
            if source == "MSCI" and score == '-':
                score = data.get("Rating", '-')
        else:
            score = '-'
            
        return score, data
    except Exception as e:
        print(f"Error fetching {source} score for {name}: {e}")
        return '-', {}

def fetch_scores(company_name, source_specific_names):
    """
    Fetch ESG scores from various sources using concurrent processing
    """
    scores = {}
    details = {}

    def get_safe_name(source):
        return source_specific_names.get(source, company_name).strip()

    # Create a list of futures for concurrent processing
    futures = []
    with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
        # Submit all API calls concurrently
        for source in ["S&P", "Sustainalytics", "ISS", "LSEG", "MSCI"]:
            name = get_safe_name(source)
            future = executor.submit(fetch_single_source, source, name)
            futures.append((source, future))

        # Collect results as they complete
        for source, future in futures:
            try:
                score, data = future.result()
                scores[source] = score
                details[source] = data
            except Exception as e:
                print(f"Error processing {source} for {company_name}: {e}")
                scores[source] = '-'
                details[source] = {}

    return scores, details

def initialize_columns_for_source(sheet, source):
    """
    Initialize columns for each source's detailed sheet with just company name and a placeholder for data fields.
    """
    sheet.append(["Company Name", "Data Fields"])

def append_details_to_sheets(details_sheets, company_name, details):
    """
    Append detailed data to respective sheets using a fully dynamic approach like the older version.
    """
    for source, data in details.items():
        source_sheet = details_sheets.get(source)
        if not source_sheet:
            continue

        # Handle ISS data which comes as a list
        if source == "ISS" and isinstance(data, list) and len(data) > 0:
            data = data[0]  # Get first item if it's a list
            
        # Use a fully dynamic approach to save all data
        if isinstance(data, dict) and data:
            # If the data is empty or just contains error info, skip it
            if len(data) <= 1 and ('error' in data or not data):
                continue
            
            # If this is the first company, set up the headers based on the data fields
            if source_sheet.max_row == 1:  # Only headers exist
                # Create dynamic headers based on the data
                source_sheet.delete_rows(1)  # Remove the initial placeholder row
                source_sheet.append(["Company Name"] + list(data.keys()))
                # Apply formatting to header
                for cell in source_sheet[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            
            # Add the row with all data fields
            values = [company_name]
            # Add each value in the correct order based on the headers
            headers = []
            for cell in source_sheet[1][1:]:  # Skip the Company Name column
                headers.append(cell.value)
                
            for header in headers:
                value = data.get(header, "N/A")
                # Convert complex values to strings
                if not isinstance(value, (str, int, float, bool, type(None))):
                    value = str(value)
                values.append(value)
            
            source_sheet.append(values)

def update_excel_file(input_file, companies_data, output_file=None):
    """
    Update an Excel file with ESG scores for companies
    """
    # Extract company data for processing
    companies = parse_excel_for_companies(input_file)
    
    # Process companies in batches for better performance
    batch_size = 10
    results = []
    
    for i in range(0, len(companies), batch_size):
        batch = companies[i:i + batch_size]
        batch_results = []
        
        # Process each company in the batch
        for company in batch:
            scores, details = fetch_scores(company["name"], company["source_specific_names"])
            batch_results.append((company["name"], scores, details))
        
        results.extend(batch_results)
        
        # Clear cache periodically to prevent memory issues
        if i % 50 == 0:
            cached_snp_score.cache_clear()
            cached_sustainalytics_score.cache_clear()
            cached_iss_score.cache_clear()
            cached_lseg_score.cache_clear()
            cached_msci_score.cache_clear()
    
    # Generate output filename if not provided
    if not output_file:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"updated_esg_scores_{timestamp}.xlsx"
    
    # Update Excel file with results
    try:
        workbook = load_workbook(input_file)
    except Exception as e:
        print(f"Error loading workbook: {e}")
        workbook = Workbook()
    
    # Create Summary sheet if it doesn't exist
    if "Summary" not in workbook.sheetnames:
        summary_sheet = workbook.create_sheet("Summary")
        summary_sheet.append([
            "Company Name", "S&P", "Sustainalytics", "ISS", "LSEG", "MSCI"
        ])
        # Apply formatting to header
        for cell in summary_sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    else:
        summary_sheet = workbook["Summary"]
    
    # Create detailed sheets for each source - do this only once
    details_sheets = {}
    for source in ["S&P", "Sustainalytics", "ISS", "LSEG", "MSCI"]:
        sheet_name = source
        if sheet_name not in workbook.sheetnames:
            sheet = workbook.create_sheet(sheet_name)
            initialize_columns_for_source(sheet, source)
            # Apply formatting to header
            for cell in sheet[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        else:
            sheet = workbook[sheet_name]
        details_sheets[source] = sheet
    
    # Process all data headers first to set up sheets correctly
    data_by_source = {}
    for _, _, details in results:
        for source, data in details.items():
            if source == "ISS" and isinstance(data, list) and len(data) > 0:
                data = data[0]
            if isinstance(data, dict) and data and not (len(data) <= 1 and ('error' in data or not data)):
                if source not in data_by_source:
                    data_by_source[source] = []
                data_by_source[source].append(data)
    
    # Set up headers for each sheet based on the first valid data
    for source, data_list in data_by_source.items():
        if not data_list:
            continue
        
        source_sheet = details_sheets.get(source)
        if source_sheet and source_sheet.max_row == 1 and data_list:  # Only headers exist
            # Use the first valid data to set headers
            source_sheet.delete_rows(1)  # Remove the initial placeholder row
            source_sheet.append(["Company Name"] + list(data_list[0].keys()))
            # Apply formatting to header
            for cell in source_sheet[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    # Populate sheets with results in bulk
    for company_name, scores, details in results:
        # Add to summary sheet
        summary_sheet.append([
            company_name,
            scores.get("S&P", '-'),
            scores.get("Sustainalytics", '-'),
            scores.get("ISS", '-'),
            scores.get("LSEG", '-'),
            scores.get("MSCI", '-')
        ])
        
        # Add to detailed sheets
        append_details_to_sheets(details_sheets, company_name, details)
    
    # Save workbook
    workbook.save(output_file)
    print(f"Updated Excel file saved as {output_file}")
    return output_file

def parse_excel_for_companies(file_path):
    """
    Parse an Excel file to extract company names and source-specific names
    """
    try:
        # Use pandas optimizations for faster reading
        df = pd.read_excel(file_path, engine='openpyxl').fillna("")
        
        # Look for company name column - optimize this search
        company_cols = [col for col in df.columns if isinstance(col, str) and "company" in col.lower() and "name" in col.lower()]
        company_col = next((col for col in company_cols if not any(s.lower() in col.lower() for s in ["S&P", "Sustainalytics", "ISS", "LSEG", "MSCI"])), None)
        
        # If not found, use the default name
        if not company_col and "Company Name" in df.columns:
            company_col = "Company Name"
            
        if not company_col:
            # Last resort - just use the first column
            company_col = df.columns[0]
            print(f"Warning: Using column '{company_col}' as company name column")
        
        # Find source-specific name columns more efficiently
        source_columns = {}
        source_keywords = ["S&P", "Sustainalytics", "ISS", "LSEG", "MSCI"]
        
        for col in df.columns:
            if not isinstance(col, str):
                continue
                
            col_lower = col.lower()
            for source in source_keywords:
                if source.lower() in col_lower and ("name" in col_lower or "company" in col_lower):
                    source_columns[source] = col
                    break
        
        # Process companies in batches for efficiency
        companies = []
        for _, row in df.iterrows():
            company_name = str(row[company_col]).strip()
            if not company_name:
                continue
                
            # Collect source-specific names with less string operations
            source_specific_names = {
                source: str(row[col]).strip() 
                for source, col in source_columns.items() 
                if row[col] and str(row[col]).strip()
            }
            
            companies.append({"name": company_name, "source_specific_names": source_specific_names})
        
        return companies
    
    except Exception as e:
        print(f"Error parsing Excel file: {e}")
        raise ValueError(f"Could not parse Excel file: {e}")

def process_company(company_data):
    """
    Process a company to get ESG scores from various sources
    
    Args:
        company_data (dict): Dictionary containing company name and source-specific names
    
    Returns:
        dict: Dictionary with company name, scores, and details
    """
    company_name = company_data["name"]
    source_specific_names = company_data.get("source_specific_names", {})
    print(f"Processing company: {company_name}")
    
    # Get the scores and details
    scores, details = fetch_scores(company_name, source_specific_names)
    
    # Return in the expected format
    return {"name": company_name, "scores": scores, "details": details}
